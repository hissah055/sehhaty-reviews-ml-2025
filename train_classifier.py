# ======================================================
# Reviews Classification — Few-shot (≤1000) — AR/EN (Merged, fixed fallback)
# ------------------------------------------------------
# - Works on a merged AR+EN file with a unified text column (🧹Content_Clean).
# - Few-shot: AR=1000 , EN=500 if Language column exists.
# - Subtheme: LinearSVC | Sentiment: LogisticRegression(balanced).
# - Theme derived from Subtheme via fixed taxonomy.
# - Sentiment falls back to Rating only when text is invalid or empty.
# - Writes predictions only to empty cells; preserves GT if present.
# - Generates summary cards for Theme / Subtheme / Sentiment.
# ======================================================

import os, re, shutil, zipfile, numpy as np, pandas as pd, unicodedata
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import LinearSVC
from sklearn.linear_model import LogisticRegression
from openpyxl import load_workbook
from collections import Counter
from IPython.display import HTML, display

# ---------- Input ----------
INPUT_XLSX = r"/content/Database_Merged_AR_EN (sample24,000).xlsx"
SHEET_NAME = None
RANDOM_SEED = 13
np.random.seed(RANDOM_SEED)

TRAIN_LIMIT    = 1000
TRAIN_LIMIT_AR = 1000
TRAIN_LIMIT_EN = 500

# ---------- Helpers ----------
def assert_xlsx_ok(path):
    if not os.path.exists(path): raise FileNotFoundError(path)
    if not zipfile.is_zipfile(path): raise ValueError("BadZipFile")

def smart_out_path(p):
    base, ext = os.path.splitext(p)
    stem = os.path.basename(base)
    new = re.sub(r'(?i)(Arabic|English)', r'\1_classification', stem, count=1)
    if new == stem: new = stem + "_classification"
    out = os.path.join(os.path.dirname(base), new + ext)
    if os.path.abspath(out) == os.path.abspath(p):
        out = os.path.join(os.path.dirname(base), new + "_classified" + ext)
    return out

def find_col(df, cands):
    cols = list(df.columns)
    low  = {str(c).strip().lower(): c for c in cols}
    for cand in cands:
        k = str(cand).strip().lower()
        if k in low: return low[k]
    comp = [re.sub(r"\s+","", str(c).strip().lower()) for c in cands]
    for c in cols:
        cc = re.sub(r"\s+","", str(c).strip().lower())
        if any(k in cc for k in comp): return c
    return None

def rating_to_sentiment(r):
    try: r = float(r)
    except: return None
    if r >= 4: return "Positive"
    if r <= 2: return "Negative"
    return "Neutral"

def is_meaningful_text(x):
    t = str(x or "").strip()
    if not t or t.lower()=="nan": return False
    return bool(re.search(r"[A-Za-z0-9\u0600-\u06FF]", t))

TOKEN_PATTERN = r"(?u)(?:[\w\u0600-\u06FF]+|[\U0001F300-\U0001FAFF\u2600-\u26FF])"

def fit_svm(X, y):
    if len(set(y)) < 2: return (None, None)
    vec = TfidfVectorizer(ngram_range=(1,2), token_pattern=TOKEN_PATTERN)
    Xv  = vec.fit_transform(list(map(str, X)))
    clf = LinearSVC(random_state=RANDOM_SEED).fit(Xv, list(map(str, y)))
    return vec, clf

def fit_lr(X, y):
    if len(set(y)) < 2: return (None, None)
    vec = TfidfVectorizer(ngram_range=(1,2), token_pattern=TOKEN_PATTERN)
    Xv  = vec.fit_transform(list(map(str, X)))
    clf = LogisticRegression(max_iter=300, class_weight="balanced",
                             random_state=RANDOM_SEED).fit(Xv, list(map(str, y)))
    return vec, clf

taxonomy = {
 "User Experience & Sentiment": ["Ease of Use","Navigation","UI Clarity","Onboarding",
   "Overall Satisfaction","Accessibility","Help & Guidance","General","General_UX"],
 "Technical Performance": ["App Speed","Loading Time","Crashes / Freezes","Errors / Bugs",
   "Connectivity / Network","Stability","General","General_Technical"],
 "Content & Services": ["Appointment Booking","Results Delivery","Reports / Documents",
   "Prescriptions","Records / Vaccination","Teleconsultation","General","General_Content"],
 "Security & Support": ["Login / OTP","Password Reset","Account Verification","Privacy / Permissions",
   "Support Responsiveness","Account Access Issues","General","General_Security"],
 "Suggestions & UI Design": ["Feature Request – Dark Mode","Notifications & Reminders","Layout Improvements",
   "Customization","Language Options","Accessibility Enhancements","General","General_Suggestions"]
}
sub_to_theme = {s: t for t, subs in taxonomy.items() for s in subs}

# ---------- Load ----------
assert_xlsx_ok(INPUT_XLSX)
OUT_XLSX = smart_out_path(INPUT_XLSX)
df = pd.read_excel(INPUT_XLSX, sheet_name=0 if SHEET_NAME is None else SHEET_NAME)

TEXT_CANDS = ["🧹Content_Clean","Content","text","comment","review"]
SUB_GT_CANDS  = ["Subtheme_GT"]
SENT_GT_CANDS = ["Sentiment_GT"]
RATING_CANDS  = ["⭐Rating","⭐ Rating","Rating","Stars","Score","التقييم"]
LANG_CANDS    = ["Language","lang"]

TEXT_COL = find_col(df, TEXT_CANDS)
SUB_GT   = find_col(df, SUB_GT_CANDS)
SENT_GT  = find_col(df, SENT_GT_CANDS)
LANG_COL = find_col(df, LANG_CANDS)
RATING_COL = find_col(df, RATING_CANDS)
ratings = df[RATING_COL] if RATING_COL else None

# ---------- Few-shot sampling ----------
def normalize_lang(series):
    s = series.astype(str).str.strip().str.upper()
    mapping = {"AR":"AR","EN":"EN","ARABIC":"AR","ENGLISH":"EN"}
    return s.replace(mapping)

def take_capped(label_col):
    base = df[[TEXT_COL, label_col] + ([LANG_COL] if LANG_COL else [])].copy()
    base = base[base[TEXT_COL].notna() & base[label_col].notna()]
    base[TEXT_COL]  = base[TEXT_COL].astype(str).str.strip()
    base[label_col] = base[label_col].astype(str).str.strip()
    base = base[(base[TEXT_COL]!="") & (base[label_col]!="")]
    if LANG_COL:
        base[LANG_COL] = normalize_lang(base[LANG_COL])
        base = base.sample(frac=1, random_state=RANDOM_SEED)
        ar = base[base[LANG_COL]=="AR"].head(TRAIN_LIMIT_AR)
        en = base[base[LANG_COL]=="EN"].head(TRAIN_LIMIT_EN)
        return pd.concat([ar, en], ignore_index=True)
    else:
        return base.sample(frac=1, random_state=RANDOM_SEED).head(TRAIN_LIMIT)

gt_sub  = take_capped(SUB_GT)
gt_sent = take_capped(SENT_GT)

vec_sub,  clf_sub  = fit_svm(gt_sub[TEXT_COL],  gt_sub[SUB_GT])   if len(gt_sub)  else (None, None)
vec_sent, clf_sent = fit_lr (gt_sent[TEXT_COL], gt_sent[SENT_GT]) if len(gt_sent) else (None, None)

print(f"📚 Few-shot sizes → Sub:{len(gt_sub)} | Sent:{len(gt_sent)}")

# ---------- Predict ----------
texts = df[TEXT_COL].astype(str).fillna("").tolist()

def predict_subtheme(txts):
    return clf_sub.predict(vec_sub.transform(txts)).tolist() if clf_sub else [""]*len(txts)

def predict_sentiment(txts):
    base = (clf_sent.predict(vec_sent.transform(txts)).tolist()
            if clf_sent else [""]*len(txts))
    out=[]
    for raw, b, r in zip(txts, base, (ratings if ratings is not None else [None]*len(txts))):
        pred_ok = str(b).strip() != ""
        if (not is_meaningful_text(raw)) or (not pred_ok):
            fb = rating_to_sentiment(r)
            out.append(fb if fb else b)
        else:
            out.append(b)
    return out

sub_pred   = predict_subtheme(texts)
sent_pred  = predict_sentiment(texts)

def map_theme(sub_val):
    if not is_meaningful_text(sub_val): return ""
    return sub_to_theme.get(str(sub_val).strip(), "User Experience & Sentiment")

# ---------- Save ----------
THEME_OUT_CANDS = ["🎯 Theme","Theme"]
SUB_OUT_CANDS   = ["🧩 Subtheme","Subtheme"]
SENT_OUT_CANDS  = ["😊 Sentiment","Sentiment"]

def pick_header_index(ws, cands):
    headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
    lower   = [str(h).strip().lower() for h in headers]
    for cand in cands:
        key = str(cand).strip().lower()
        for i, h in enumerate(lower):
            if key == h: return i+1
    comp = [re.sub(r"\s+","", str(c).strip().lower()) for c in cands]
    for i, h in enumerate(lower):
        if any(k in h for k in comp): return i+1
    return None

shutil.copyfile(INPUT_XLSX, OUT_XLSX)
wb = load_workbook(OUT_XLSX)
ws = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]

c_theme = pick_header_index(ws, THEME_OUT_CANDS)
c_sub   = pick_header_index(ws, SUB_OUT_CANDS)
c_sent  = pick_header_index(ws, SENT_OUT_CANDS)
gt_sub_avail  = df[SUB_GT ].notna() & (df[SUB_GT ].astype(str).str.strip()!="")
gt_sent_avail = df[SENT_GT].notna() & (df[SENT_GT].astype(str).str.strip()!="")

fill_counts = {"theme":0,"sub":0,"sent":0}
n_rows = min(len(df), ws.max_row-1)

for i in range(n_rows):
    r = i + 2

    if c_sub and str(ws.cell(row=r, column=c_sub).value or "").strip()=="":
        if gt_sub_avail.iloc[i]:
            ws.cell(row=r, column=c_sub, value=str(df.at[i, SUB_GT]).strip()); fill_counts["sub"] += 1
        else:
            val = str(sub_pred[i]).strip()
            if val:
                ws.cell(row=r, column=c_sub, value=val); fill_counts["sub"] += 1

    if c_theme and str(ws.cell(row=r, column=c_theme).value or "").strip()=="":
        base_sub = str(df.at[i, SUB_GT]).strip() if gt_sub_avail.iloc[i] else str(sub_pred[i]).strip()
        t = map_theme(base_sub)
        if t:
            ws.cell(row=r, column=c_theme, value=t); fill_counts["theme"] += 1

    if c_sent and str(ws.cell(row=r, column=c_sent).value or "").strip()=="":
        if gt_sent_avail.iloc[i]:
            ws.cell(row=r, column=c_sent, value=str(df.at[i, SENT_GT]).strip()); fill_counts["sent"] += 1
        else:
            val = str(sent_pred[i]).strip()
            if val:
                ws.cell(row=r, column=c_sent, value=val); fill_counts["sent"] += 1

wb.save(OUT_XLSX)
print("✅ Saved:", OUT_XLSX)
print(f"Filled → Theme:{fill_counts['theme']} | Subtheme:{fill_counts['sub']} | Sentiment:{fill_counts['sent']}")

# ---------- Summary Cards ----------
def fancy_card(title, items, emoji):
    c = Counter([x for x in items if x and str(x).strip()!=""])
    rows = "".join(
        f"<tr><td style='padding:4px 8px'>{k}</td>"
        f"<td style='text-align:right;padding:4px 8px'>{v}</td></tr>"
        for k,v in c.most_common()
    )
    return f"""
    <div style="max-width:560px;width:100%;margin:10px 0;
                border:1px solid #90caf9;border-radius:14px;
                background:linear-gradient(135deg,#e3f2fd 0%, #fff7e6 100%);
                padding:12px;box-shadow:0 2px 6px rgba(0,0,0,.08)">
      <div style="font-weight:800;margin-bottom:6px;display:flex;gap:8px;align-items:center">
        <span style="font-size:22px;color:#0d47a1">{emoji}</span>
        <span style="color:#0d47a1">{title}</span>
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:12px">
        <thead>
          <tr>
            <th style="text-align:left;padding:4px 8px;color:#0d47a1;background:#dbeafe">Class</th>
            <th style="text-align:right;padding:4px 8px;color:#7a4a00;background:#ffedd5">Count</th>
          </tr>
        </thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""

theme_for_cards = [sub_to_theme.get(str(x).strip(), "User Experience & Sentiment")
                   for x in sub_pred if x and str(x).strip()!=""]

display(HTML(
    fancy_card("Theme Predictions",    theme_for_cards, "🏷️") +
    fancy_card("Subtheme Predictions", sub_pred,        "🧩") +
    fancy_card("Sentiment Predictions",sent_pred,       "🙂")
))
